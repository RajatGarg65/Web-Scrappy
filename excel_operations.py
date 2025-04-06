import pandas as pd
from logging_config import logger
from openpyxl import Workbook, load_workbook
import os
import json

def read_input_file(filename):
    """
    Reads an input file (Excel or CSV) and returns a DataFrame containing the 'parent_url' column.

    Args:
        filename (str): The path to the input file.

    Returns:
        pd.DataFrame: A DataFrame with the 'parent_url' column if successful; otherwise, None.
    """
    # Get file extension

    _, file_extension = os.path.splitext(filename)
    # Check if the file is Excel

    if file_extension.lower() in ['.xlsx', '.xls']:
        try:
            # Read Excel file

            return pd.read_excel(filename, usecols=['parent_url'])
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            print(f"Error reading Excel file: {e}")
            print("Trying to read as CSV...")
    # Attempt to read CSV file if Excel reading failed

    try:
        return pd.read_csv(filename, usecols=['parent_url'])
    except Exception as e:
        logger.error(f"Error reading CSV file: {e}")
        print(f"Error reading CSV file: {e}")
        return None

def get_excel_file_path(parent_url):
    """
    Generates a file path for an Excel file based on the provided parent URL.

    Args:
        parent_url (str): The parent URL used to create a safe filename.

    Returns:
        str: The generated file path for the Excel file.
    """
    # Create a safe filename from the parent URL
    safe_filename = "".join([c for c in parent_url if c.isalpha() or c.isdigit() or c==' ']).rstrip()
    output_dir = os.path.join(os.path.dirname(__file__), 'outputs')
    os.makedirs(output_dir, exist_ok=True)
    return os.path.join(output_dir, f'output_{safe_filename}.xlsx')


def update_excel(parent_url, row, pagination_info):
    """
    Updates or creates an Excel file with new data.

    Args:
        parent_url (str): The parent URL used to determine the file path.
        row (list): The data row to append to the Excel file.
        pagination_info (dict): Pagination information to update in the row if applicable.

    Returns:
        None
    """
    excel_file = get_excel_file_path(parent_url)
    # Create a new Excel file if it doesn't exist

    if not os.path.exists(excel_file):
        logger.info(f"Excel file not found. Creating new workbook: {excel_file}")
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Parent_url', 'Extracted links', 'Potential release', 'Final releases', 
                   'Pagination parent url', 'Pagination links', 'Number of pages'])
        try:
            wb.save(excel_file)
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return
    # Load the existing Excel file

    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        return
    # Get the potential release URL from the row

    potential_release_url = row[3]
    logger.info(f"Potential Release URL: {potential_release_url}")
    # Check if there is matching pagination info

    matching_pagination = next((
        (parent_url, info) for parent_url, info in pagination_info.items()
        if parent_url == potential_release_url
    ), None)
    # Update the row with pagination info if found

    if matching_pagination:
        parent_url, page_info = matching_pagination
        logger.info(f"Found pagination info for {parent_url}: {page_info}")
        row[5] = parent_url
        row[6] = json.dumps(page_info['pagination_links'])
        row[7] = page_info['page_count']
    else:
        logger.info(f"No pagination info found for {potential_release_url}")
        row[5] = ''
        row[6] = '[]'
        row[7] = ''
    # Append the updated row to the sheet

    ws.append(row)
    logger.info(f"Row appended: {row}")
    # Save changes to the Excel file

    try:
        wb.save(excel_file)
        logger.info(f"Excel file saved successfully: {excel_file}")
    except PermissionError:
        logger.error(f"Unable to save Excel file. It might be open in another program.")
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")

def remove_row_from_excel(parent_url, url):
    """
    Removes a row from the Excel file where the URL matches the given URL.

    Args:
        parent_url (str): The parent URL used to determine the file path.
        url (str): The URL to identify the row to be removed.

    Returns:
        None
    """
    excel_file = get_excel_file_path(parent_url)
    # Create a new Excel file if it doesn't exist

    if not os.path.exists(excel_file):
        logger.warning(f"Excel file not found: {excel_file}. Creating a new file.")
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Parent_url', 'Extracted links', 'Potential release', 'Final releases', 
                   'Pagination parent url', 'Pagination links', 'Number of pages'])
        wb.save(excel_file)
    # Load the existing Excel file

    
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        rows_to_delete = []
        # Identify rows to delete

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[3] == url:
                rows_to_delete.append(idx)
        # Delete identified rows

        for row_id in reversed(rows_to_delete):
            ws.delete_rows(row_id)
        # Save the changes to the Excel file
 
        if rows_to_delete:
            wb.save(excel_file)
            print(f"Removed row with URL: {url} from Excel file.")
        else:
            print(f"No matching row found for URL: {url} in Excel file.")
    except Exception as e:
        logger.error(f"Error manipulating Excel file: {e}")